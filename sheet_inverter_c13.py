import openpyxl
from pathlib import Path

# A function to invert the rows and columns of a targetFile and return an
#inverted file
def inverter(targetFile, invertedFile):
    wb = openpyxl.load_workbook(Path.home()/'wilsonian'/targetFile) # The target
    # workbook object
    sheet = wb.active # The target sheet object
    newWb = openpyxl.Workbook() # The inverted workbook Object
    newSheet = newWb.active # The inverted sheet Object
    for x, column in enumerate(list(sheet.columns)): # for loop to iterate over
    # the columns of the target sheet with x depicting the column number
        for y, cellObj in enumerate(column): # for loop iterating over the cell
        # in the column with y depicting the row number
             newSheet.cell(row = x + 1, column = y + 1).value = cellObj.value #
             # Assigning each column value to a new row by swapping the column
             # number with the row number
    newWb.save(Path.home()/'wilsonian'/invertedFile)
